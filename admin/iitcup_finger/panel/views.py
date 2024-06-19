import csv
import json
from multiprocessing import AuthenticationError
from django.shortcuts import render, redirect
from .models import Casos, Huellas, Implicado, Personas
from django.shortcuts import get_object_or_404
from django.core.files.base import ContentFile
from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import Group, User
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from io import StringIO
from django.http import JsonResponse
from django.http import HttpResponse
import numpy as np
from PIL import Image

from io import BytesIO
from io import StringIO
import io
import base64
from skimage.feature import hog
from skimage import exposure
import math
import sys
import cv2
from scipy import signal
from scipy import ndimage
import scipy
import cv2
import numpy as np
import skimage.morphology
from skimage.morphology import convex_hull_image, erosion
from skimage.morphology import square
import math
from django.core.paginator import Paginator
import openpyxl
import os
import numpy as np
import skimage
from skimage.morphology import skeletonize, convex_hull_image, erosion, square

import skimage.draw
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

TEMPLATES_DIRS={
    'os.path.join(BASE_DIR,"templates")'
}

def index(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('inicio')  
    else:
        form = AuthenticationError()
    return render(request, 'index.html', {'form': form})

def registrar(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
  
            user = form.save()


            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)


            auxiliares_group = Group.objects.get(name='Auxiliares')  # Asegúrate de que el grupo "Auxiliares" exista
            auxiliares_group.user_set.add(user)


            login(request, user)  
            

            return redirect('index')
        else:
            messages.error(request, "La contraseña no es lo suficientemente segura. Debe tener al menos 8 caracteres y contener letras mayúsculas, minúsculas, números y caracteres especiales.")
    else:
        form = UserCreationForm()
    
    return render(request, 'registrar.html', {'form': form})

admin_group, _ = Group.objects.get_or_create(name='Administradores')
for perm in Permission.objects.all():
    admin_group.permissions.add(perm)

perito_group, _ = Group.objects.get_or_create(name='Peritos') 
perms_peritos = Permission.objects.filter(codename__in=['add_caso', 'change_caso'])
perito_group.permissions.set(perms_peritos)
auxiliar_group, _ = Group.objects.get_or_create(name='Auxiliares')
perms_auxiliares = Permission.objects.filter(codename__in=['view_caso'])
auxiliar_group.permissions.set(perms_auxiliares)
admin_group.save()
perito_group.save()
auxiliar_group.save()

def listar_rol(request):
    if request.method == 'POST':
        user_id = request.POST.get('username') 
        role_name = request.POST.get('role')  
        try:
            user = User.objects.get(pk=user_id)
            group = Group.objects.get(name=role_name)
            user.groups.clear()  
            user.groups.add(group)
            return redirect('listar')
        except User.DoesNotExist:
            return HttpResponse('Usuario no encontrado.', status=404)
        except Group.DoesNotExist:
            return HttpResponse('Grupo no encontrado.', status=404)
        except Exception as e:
            return HttpResponse(f'Error al asignar el rol: {str(e)}', status=500)
    else:
        users = User.objects.all()  # Obtén todos los usuarios
        roles = Group.objects.all()  # Obtén todos los roles disponibles
        context = {'users': users, 'roles': roles}
        return render(request, 'usuarios/listar_rol.html', context)

def asignar_grupo(request):
    username = request.POST.get('username')
    group_id = request.POST.get('group_id')
    
    try:
        user = User.objects.get(username=username)
        group = Group.objects.get(id=group_id)
        user.groups.add(group)
        user.save()
        return JsonResponse({"success": "Grupo asignado correctamente."})
    except User.DoesNotExist:
        return JsonResponse({"error": "Usuario no encontrado."}, status=404)
    except Group.DoesNotExist:
        return JsonResponse({"error": "Grupo no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def listar(request):
    users = User.objects.all() 
    groups = Group.objects.all()
    return render(request, 'usuarios/listar.html', {'users': users, 'groups': groups})
    
def agregar(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')

        new_user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name
        )


        auxiliares_group = Group.objects.get(name='Auxiliares')  
        auxiliares_group.user_set.add(new_user)

        return redirect('listar')
    else:
        return render(request, "usuarios/agregar.html")

def eliminar(request):
    return render(request, "usuarios/eliminar.html")

def editar(request):
    return render(request, "usuarios/editar.html")

def inicio(request):
    return render(request, "inicio.html")


def listar_huella(request):
    huellas_list = Huellas.objects.all().order_by('-id_huella')  # Ordenar las huellas por ID de manera descendente
    paginator = Paginator(huellas_list, 7)  # Mostrar 7 huellas por página

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "huellas/listar_huella.html", {'page_obj': page_obj})



def obtener_detalle_caso(request):
    codigo_caso = request.GET.get('codigo_caso')
    if codigo_caso:
        try:
            caso = Casos.objects.get(codigo_caso=codigo_caso, activate=True)
            data = {
                'codigo_caso': caso.codigo_caso,
                'nombre_caso': caso.nombre_caso,
                'descripcion': caso.descripcion,
                'departamento': caso.departamento,
            }
            return JsonResponse(data)
        except Casos.DoesNotExist:
            return JsonResponse({'error': 'Caso no encontrado'}, status=404)
    return JsonResponse({'error': 'Código del caso no proporcionado'}, status=400)

    huella = request.FILES.get('file')
    if huella:
        huella_instancia = Huellas()
        huella_instancia.caso = idcaso
        huella_instancia.original_image.save(huella.name, huella, save=False)
  

import cv2
import numpy as np
from django.core.files.base import ContentFile
from io import BytesIO
from .models import Huellas
from django.shortcuts import render, redirect


def agregar_huella(request):
    casos_activos = Casos.objects.filter(activate=True).order_by('-f_registro')
    
    if request.method == "POST":
        idcaso = request.POST.get('codigo_caso')
        huella = request.FILES.get('file')
        
        if huella:
            try:
                nparr = np.fromstring(huella.read(), np.uint8)
                img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)
                procesada_img = PreprocesamientoHuellas().mejorar(img)
                procesada_img = (procesada_img * 255).astype(np.uint8)
                procesada_img2 = 255-procesada_img
                datos,imagen_con_minucias =  extraer_caracteristicas_minuciosas(procesada_img)
                terminaciones = [(x, y) for x, y in zip(datos['terminoX'], datos['terminoY'])]
                bifurcaciones = [(x, y) for x, y in zip(datos['bifurX'], datos['bifurY'])]
                coordenadas_minucias = terminaciones + bifurcaciones
                imagen_con_minucias=255-imagen_con_minucias

                classification = clasificar_huella(imagen_con_minucias)
                
                imag = cv2.cvtColor(imagen_con_minucias, cv2.COLOR_BGR2GRAY)
                nucleus_coords = buscar_nucleo_esqueleto(imag)


                huella_instancia = Huellas()
                huella_instancia.caso = idcaso
                huella_instancia.original_image.save(huella.name, huella, save=False)
                huella_instancia.procesada_image.save(huella.name, ContentFile(cv2.imencode('.png', procesada_img2)[1]), save=False)
                
                huella_instancia.tipo = classification 
                huella_instancia.minutiae_image.save(huella.name, ContentFile(cv2.imencode('.png', imagen_con_minucias)[1]), save=False)
                huella_instancia.save()


                if nucleus_coords:
                    csv_bytes_io = BytesIO()
                    csv_filename = f"{huella.name.split('.')[0]}_minucias.csv"
                    csv_string_io = StringIO()
                    writer = csv.writer(csv_string_io)
                    writer.writerow(["X", "Y", "Distancia", "Angulo"])
                    for point in coordenadas_minucias:
                        distance, angle = calculate_distance_angle(nucleus_coords, point)
                        writer.writerow([point[0], point[1], distance, angle])
                    csv_string_io.seek(0)
                    huella_instancia.minutiae_csv.save(csv_filename, ContentFile(csv_string_io.getvalue().encode('utf-8')), save=True)
                    huella_instancia.save()

                return redirect('listar_huella')
            
            except Exception as e:
                print(f"Error al procesar la imagen: {str(e)}")
                return render(request, "huellas/agregar_huella.html", {'casos_activos': casos_activos, 'error_message': 'Error al procesar la imagen'})
        
        else:
            return redirect('listar_huella')

    else:
        datos = {'r2': 'No se puede procesar!!'}
    
    return render(request, "huellas/agregar_huella.html", {'casos_activos': casos_activos})


def eliminar_huella(request):
    return render(request, "huellas/eliminar_huella.html")

def editar_huella(request):
    return render(request, "huellas/editar_huella.html")

def listar_caso(request):
    casos= Casos.objects.all()
    datos = {'casos': casos}
    return render(request, "casos/listar_caso.html", datos)

def listar_implicado(request):
    implicados= Implicado.objects.all()
    datos = {'implicados': implicados}
    return render(request, "casos/listar_implicado.html", datos)

def agregar_caso(request):
    if request.method == "POST":
        if request.POST.get("nombre_caso") and request.POST.get("descripcion") and request.POST.get("codigo_caso") and request.POST.get("departamento"):
            caso = Casos()
            caso.codigo_caso = request.POST.get("codigo_caso")
            caso.nombre_caso = request.POST.get("nombre_caso")
            caso.descripcion = request.POST.get("descripcion")
            caso.departamento = request.POST.get("departamento")
            caso.save()
            return redirect('listar_caso')
    else:
        return render(request, "casos/agregar_caso.html")

def eliminar_caso(request):
    return render(request, "casos/eliminar_caso.html")

def agregar_implicado(request):
    return render(request, "casos/agregar_implicado.html")

def form_editar_caso(request, id_caso):
    if request.method == "POST":
        id_caso = request.POST.get("id_caso")
        caso = get_object_or_404(Casos, id_caso=id_caso)
        caso.codigo_caso = request.POST.get("codigo_caso", caso.codigo_caso)
        caso.nombre_caso = request.POST.get("nombre_caso", caso.nombre_caso)
        caso.descripcion = request.POST.get("descripcion", caso.descripcion)
        caso.departamento = request.POST.get("departamento", caso.departamento)
        caso.save()
        return JsonResponse({'success': True})
    else:
        casos = Casos.objects.all()
        datos = {'casos': casos}
        return render(request, "casos/listar_caso.html", datos)
    
def obtener_datos_caso(request, id_caso):
    caso = Casos.objects.get(pk=id_caso)
    datos = {
        'id_caso': caso.id_caso,
        'codigo_caso': caso.codigo_caso,
        'nombre_caso': caso.nombre_caso,
        'descripcion': caso.descripcion,
        'f_registro': caso.f_registro.strftime('%Y-%m-%d'),
        'f_modificacion': caso.f_modificacion.strftime('%Y-%m-%d'),
        'departamento': caso.departamento,
    }
    return JsonResponse(datos)

def cambiar_estado_caso(request, id_caso):
    try:
        data = json.loads(request.body)
        nuevo_estado = data.get('activate')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    caso = get_object_or_404(Casos, pk=id_caso)  
    caso.activate = nuevo_estado
    caso.save()
    return JsonResponse({'success': True, 'activate': caso.activate})

def cambiar_estado_huella(request, id_caso):
    try:
        data = json.loads(request.body)
        nuevo_estado = data.get('activate')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    caso = get_object_or_404(Huellas, pk=id_caso) 
    caso.activate = nuevo_estado
    caso.save()
    return JsonResponse({'success': True, 'activate': caso.activate})


################# ALGORITMO DE RECONOCIMIENTO #######################
class ReconocimientoHuella(object):
    def __init__(self):
        self._mascara = []
        self._esqueleto = []
        self.minucias_termino = []
        self.minucias_bifurcacion = []
        self.caracteristicas_minuciosas = []

    def esqueleto_imagen(self, image):
        skeleton = np.zeros(image.shape, np.uint8)
        size = np.size(image)
        element = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
        done = False

        while not done:
            eroded = cv2.erode(image, element)
            temp = cv2.dilate(eroded, element)
            temp = cv2.subtract(image, temp)
            skeleton = cv2.bitwise_or(skeleton, temp)
            image = eroded.copy()
            zeros = size - cv2.countNonZero(image)
            if zeros == size:
                done = True
        return skeleton

    def esqueletizar(self, img):
        img = np.uint8(img > 128)
        self._esqueleto = self.esqueleto_imagen(img)
        self._esqueleto = skeletonize(img)
        self._esqueleto = np.uint8(self._esqueleto) * 255
        self._mascara = img * 255

    def calcularAngulo(self, bloque, tipoMinucia):
        angulo = []
        (filas, columnas) = np.shape(bloque)
        CentroX, CentroY = (filas - 1) / 2, (columnas - 1) / 2
        if tipoMinucia.lower() == 'terminacion':
            sumaVal = 0
            for i in range(filas):
                for j in range(columnas):
                    if (i == 0 or i == filas - 1 or j == 0 or j == columnas - 1) and bloque[i][j] != 0:
                        angulo.append(-math.degrees(math.atan2(i - CentroY, j - CentroX)))
                        sumaVal += 1
                        if sumaVal > 1:
                            angulo.append(float('nan'))
            return angulo
        elif tipoMinucia.lower() == 'bifurcacion':
            angulo = []
            sumaVal = 0
            for i in range(filas):
                for j in range(columnas):
                    if (i == 0 or i == filas - 1 or j == 0 or j == columnas - 1) and bloque[i][j] != 0:
                        angulo.append(-math.degrees(math.atan2(i - CentroY, j - CentroX)))
                        sumaVal += 1
            if sumaVal != 3:
                angulo.append(float('nan'))
            return angulo

    def obtener_puntos_caracteristicos(self):
        self._esqueleto = self._esqueleto == 255
        (filas, columnas) = self._esqueleto.shape
        self.minuciasTermino = np.zeros_like(self._esqueleto, dtype=bool)
        self.minuciasBifurcacion = np.zeros_like(self._esqueleto, dtype=bool)

        for i in range(1, filas - 1):
            for j in range(1, columnas - 1):
                if self._esqueleto[i][j] == 1:
                    bloque = self._esqueleto[i - 1:i + 2, j - 1:j + 2]
                    valorBloque = np.sum(bloque)
                    if valorBloque == 2:
                        self.minuciasTermino[i, j] = True
                    elif valorBloque == 4:
                        self.minuciasBifurcacion[i, j] = True

        self._mascara = convex_hull_image(self._mascara > 0)
        self._mascara = erosion(self._mascara, square(5))
        self.minuciasTermino = np.uint8(self._mascara) * self.minuciasTermino
        self.minuciasBifurcacion = np.uint8(self._mascara) * self.minuciasBifurcacion

    def eliminar_poros(self, listaMinucias, img, umbral):
        img = img * 0
        MinuciasEspurias = []
        numPuntos = len(listaMinucias)
        D = np.zeros((numPuntos, numPuntos))
        for i in range(1, numPuntos):
            for j in range(0, i):
                X1, Y1 = listaMinucias[i].centroid
                X2, Y2 = listaMinucias[j].centroid

                dist = np.sqrt((X2 - X1) ** 2 + (Y2 - Y1) ** 2)
                D[i][j] = dist
                if dist < umbral:
                    MinuciasEspurias.append(i)
                    MinuciasEspurias.append(j)

        MinuciasEspurias = np.unique(MinuciasEspurias)
        for i in range(0, numPuntos):
            if not i in MinuciasEspurias:
                X, Y = np.int16(listaMinucias[i].centroid)
                img[X, Y] = 1

        img = np.uint8(img)
        return img

    def buscar_nucleo(self):
        centro_y, centro_x = np.array(self._esqueleto.shape) // 2
        region_central = self._esqueleto[centro_y - 50:centro_y + 50, centro_x - 50:centro_x + 50]
        kernel = np.ones((5, 5), np.uint8)
        dilatado = cv2.dilate(region_central, kernel, iterations=1)
        momentos = cv2.moments(dilatado)
        if momentos["m00"] != 0:
            centroide_x = int(momentos["m10"] / momentos["m00"])
            centroide_y = int(momentos["m01"] / momentos["m00"])
            return (centro_x - 50 + centroide_x, centro_y - 50 + centroide_y)
        else:
            return None

    def buscar_delta(self):
        esquinas = cv2.goodFeaturesToTrack(self._esqueleto, 100, 0.01, 10)
        centro_x, centro_y = np.array(self._esqueleto.shape) // 2
        distancia_minima = np.inf
        delta = None
        for esquina in esquinas:
            x, y = esquina.ravel()
            distancia = np.sqrt((x - centro_x) ** 2 + (y - centro_y) ** 2)
            if distancia < distancia_minima:
                distancia_minima = distancia
                delta = (int(x), int(y))
        return delta

    def limpiar_ruido(self, img):
        self.minuciasTermino = skimage.measure.label(self.minuciasTermino, connectivity=2)
        RP = skimage.measure.regionprops(self.minuciasTermino)
        self.minuciasTermino = self.eliminar_poros(RP, np.uint8(img), 10)

    def marcar_puntos_caracteristicos(self):
        TerminoX = []
        TerminoY = []
        BifurcacionX = []
        BifurcacionY = []
        self.minuciasTermino = skimage.measure.label(self.minuciasTermino, connectivity=2)
        RP = skimage.measure.regionprops(np.uint8(self.minuciasTermino))
        TamanioVentana = 2
        for num, i in enumerate(RP):
            (fila, columna) = np.int16(np.round(i['centroid']))
            bloque = self._esqueleto[fila - TamanioVentana:fila + TamanioVentana + 1,
                     columna - TamanioVentana:columna + TamanioVentana + 1]
            angulo = self.calcularAngulo(bloque, 'Terminacion')
            if len(angulo) == 1:
                TerminoX.append(fila)
                TerminoY.append(columna)

        self.minuciasBifurcacion = skimage.measure.label(self.minuciasBifurcacion, connectivity=2)
        RP = skimage.measure.regionprops(np.uint8(self.minuciasBifurcacion))
        TamanioVentana = 1
        for i in RP:
            (fila, columna) = np.int16(np.round(i['centroid']))
            bloque = self._esqueleto[fila - TamanioVentana:fila + TamanioVentana + 1,
                     columna - TamanioVentana:columna + TamanioVentana + 1]
            angulo = self.calcularAngulo(bloque, 'Bifurcacion')
            if len(angulo) == 3:
                BifurcacionX.append(fila)
                BifurcacionY.append(columna)
        return {'terminoX': TerminoX, 'terminoY': TerminoY, 'bifurX': BifurcacionX, 'bifurY': BifurcacionY}

    def extraer_caracteristicas(self, img):
        self.esqueletizar(img)
        self.obtener_puntos_caracteristicos()
        self.limpiar_ruido(img)
        return self.marcar_puntos_caracteristicos()

    def mostrar_resultados(self):
        BifLabel = skimage.measure.label(self.minuciasBifurcacion, connectivity=2)
        TermLabel = skimage.measure.label(self.minuciasTermino, connectivity=2)

        minuciasBif = TermLabel * 0
        minuciasTerm = BifLabel * 0

        (filas, columnas) = self._esqueleto.shape
        ImgDisp = np.zeros((filas, columnas, 3), np.uint8)
        ImgDisp[:, :, 0] = 255 * self._esqueleto
        ImgDisp[:, :, 1] = 255 * self._esqueleto
        ImgDisp[:, :, 2] = 255 * self._esqueleto

        RP = skimage.measure.regionprops(BifLabel)
        for idx, i in enumerate(RP):
            (fila, columna) = np.int16(np.round(i['centroid']))
            minuciasBif[fila, columna] = 1
            (rr, cc) = skimage.draw.circle_perimeter(fila, columna, 4)
            skimage.draw.set_color(ImgDisp, (rr, cc), (0, 255, 255))

        RP = skimage.measure.regionprops(TermLabel)
        for idx, i in enumerate(RP):
            (fila, columna) = np.int16(np.round(i['centroid']))
            minuciasTerm[fila, columna] = 1
            (rr, cc) = skimage.draw.circle_perimeter(fila, columna, 4)
            skimage.draw.set_color(ImgDisp, (rr, cc), (255, 255, 0))

        return ImgDisp

def extraer_caracteristicas_minuciosas(img):
    extractor_caracteristicas = ReconocimientoHuella()
    resultados = extractor_caracteristicas.extraer_caracteristicas(img)
    img_resultado = extractor_caracteristicas.mostrar_resultados()
    
    return resultados, img_resultado


def clasificar_huella(image):
    image_blurred = cv2.GaussianBlur(image, (5, 5), 0)
    edges = cv2.Canny(image_blurred, 100, 200)
    if len(image.shape) > 2:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    fd, hog_image = hog(image, orientations=8, pixels_per_cell=(16, 16),
                        cells_per_block=(1, 1), visualize=True, feature_vector=True)
    hog_image_rescaled = exposure.rescale_intensity(hog_image, in_range=(0, 10))
    hog_sum = fd.sum()
    print(hog_sum)
    if hog_sum < 300:
        classification = 'Arco'
    elif hog_sum < 400:
        classification = 'Lazo'
    else:
        classification = 'Espiral'

    return classification

def buscar_nucleo_esqueleto(esqueleto):
    binary_image = cv2.bitwise_not(esqueleto)
    contours, _ = cv2.findContours(binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if contours:
        moments = [cv2.moments(cnt) for cnt in contours]
        filtered_moments = [moment for moment in moments if moment['m00'] != 0]
        if filtered_moments:
            areas = [moment['m00'] for moment in filtered_moments]
            max_area_index = np.argmax(areas)
            M = filtered_moments[max_area_index]
            cX = int(M['m10'] / M['m00'])
            cY = int(M['m01'] / M['m00'])
            # cv2.circle(esqueleto, (cX, cY), 25, (0, 255, 255), -1)
            return (cX, cY)  
    return None

import csv
from django.db.models import Q

def comparar_con_huellas_guardadas(minutiae_data_subida):
    resultados = []
    umbral_distancia = 5  
    umbral_angulo = 10    
    huellas = Huellas.objects.all()
    
    for huella in huellas:
        path = huella.minutiae_csv.path
        coincidencias = 0    
        try:
            with open(path, 'r', encoding='utf-8') as csvfile:
                csvreader = csv.DictReader(csvfile)
                datos_guardados = list(csvreader)               
                for dato_subida in minutiae_data_subida:
                    for dato_guardado in datos_guardados:
                        distancia_guardada = float(dato_guardado['Distancia'])
                        angulo_guardado = float(dato_guardado['Angulo'])
                        distancia_subida = float(dato_subida['distance'])  
                        angulo_subida = float(dato_subida['angle'])
                        if abs(distancia_subida - distancia_guardada) <= umbral_distancia and abs(angulo_subida - angulo_guardado) <= umbral_angulo:
                            coincidencias += 1
                            break  

            porcentaje_coincidencia = (coincidencias / len(datos_guardados)) * 100 if datos_guardados else 0
            if huella.persona:
                try:
                    implicado = Implicado.objects.get(carnet_identidad=huella.persona)
                    nombre_completo = f"{implicado.nombres} {implicado.apellido_paterno} {implicado.apellido_materno}"
                except Implicado.DoesNotExist:
                    nombre_completo = "No encontrado"
            else:
                nombre_completo = "No especificado"

            resultados.append({
                'huella_id': huella.id_huella, 
                'porcentaje_coincidencia': porcentaje_coincidencia,
                'persona': nombre_completo
            })
            
        except UnicodeDecodeError as e:
            print(f"Error al leer el archivo {path}: {e}")
    
    mejores_resultados = sorted(resultados, key=lambda x: x['porcentaje_coincidencia'], reverse=True)[:5]
    return mejores_resultados


#####################################################################

################# ALGORITMO DE PREPROCESAMIENTO #######################
class PreprocesamientoHuellas:

    def __init__(self):
        self.factor_escala_relativo_x = 0.65
        self.factor_escala_relativo_y = 0.65
        self.frecuencia = None
        self.frecuencia_media = None
        self.frecuencia_mediana = None
        self.imagen_binaria = None
        self.imagen_normalizada = None
        self.incremento_angulo = 3
        self.longitud_de_onda_maxima = 15
        self.longitud_de_onda_minima = 5
        self.mascara = None
        self.sigma_bloque = 7
        self.sigma_gradiente = 1
        self.sigma_suavizado_orientacion = 7
        self.tam_bloque_frecuencia_crestas = 38
        self.tam_bloque_segmento_crestas = 16
        self.tam_ventana_frecuencia_crestas = 5
        self.umbral_filtro_crestas = -3
        self.umbral_segmento_crestas = 0.1
        self.nuevas_filas_redimensionar = 350

    def normalizar(self, img: np.ndarray) -> np.ndarray:
        img = img.astype(np.float64)

        # Crear un archivo Excel con la imagen original
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, fila in enumerate(img, start=1):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)
        wb.save(os.path.join('resultados_algoritmos', 'imagen_original.xlsx'))

        if len(img.shape) == 3: 
            for c in range(img.shape[2]):
                canal = img[:, :, c]
                promedio_canal = np.mean(canal)
                varianza_canal = np.var(canal)
                if varianza_canal == 0:
                    raise ValueError("Revisa la imagen")
                img[:, :, c] = (canal - promedio_canal) / np.sqrt(varianza_canal)
        else:
            promedio_imagen_pixeles = np.mean(img)
            varianza_imagen_pixeles = np.var(img)
            if varianza_imagen_pixeles == 0:
                raise ValueError("La varianza de la imagen es cero - revise la imagen cumpla los parametros")

            img = (img - promedio_imagen_pixeles) / np.sqrt(varianza_imagen_pixeles)

        # Crear un archivo Excel con la imagen normalizada
        for i, fila in enumerate(img, start=1):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)
        wb.save(os.path.join('resultados_algoritmos', 'imagen_normalizada.xlsx'))

        # Normalizar la imagen para que los valores estén en el rango [0, 255]
        img_normalizada = (img * 255).astype(np.uint8)

        # Guardar la imagen normalizada en formato JPG utilizando OpenCV
        cv2.imwrite(os.path.join('resultados_algoritmos', 'imagen_normalizada.jpg'), img_normalizada)

        return img
    
    def segmentar_crestas(self, imagen_normalizada: np.ndarray):
            filas, columnas = imagen_normalizada.shape
            nuevas_filas = int(self.tam_bloque_segmento_crestas * np.ceil(filas / self.tam_bloque_segmento_crestas))
            nuevas_columnas = int(self.tam_bloque_segmento_crestas * np.ceil(columnas / self.tam_bloque_segmento_crestas))
            imagen_rellenada = np.zeros((nuevas_filas, nuevas_columnas))
            desviacion_estandar_imagen = np.zeros((nuevas_filas, nuevas_columnas))
            imagen_rellenada[:filas, :columnas] = imagen_normalizada

            # Guardar imagen rellenada en un archivo Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            for i, fila in enumerate(imagen_rellenada, start=1):
                for j, valor in enumerate(fila, start=1):
                    ws.cell(row=i, column=j, value=valor)
            wb.save(os.path.join('resultados_algoritmos', 'imagen_segmentada_1.xlsx'))

            for i in range(0, nuevas_filas, self.tam_bloque_segmento_crestas):
                for j in range(0, nuevas_columnas, self.tam_bloque_segmento_crestas):
                    bloque = imagen_rellenada[i:i + self.tam_bloque_segmento_crestas, j:j + self.tam_bloque_segmento_crestas]
                    desviacion_estandar_imagen[i:i + self.tam_bloque_segmento_crestas, j:j + self.tam_bloque_segmento_crestas] = np.std(bloque)

            desviacion_estandar_imagen = desviacion_estandar_imagen[:filas, :columnas]
            self.mascara = desviacion_estandar_imagen > self.umbral_segmento_crestas
            valor_medio = np.mean(imagen_normalizada[self.mascara])
            valor_desviacion_estandar = np.std(imagen_normalizada[self.mascara])
            self.imagen_normalizada = (imagen_normalizada - valor_medio) / valor_desviacion_estandar

            # Guardar imagen normalizada en un archivo Excel
            for i, fila in enumerate(self.imagen_normalizada, start=1):
                for j, valor in enumerate(fila, start=1):
                    ws.cell(row=i, column=j, value=valor)
            wb.save(os.path.join('resultados_algoritmos', 'imagen_segmentada_2.xlsx'))

            # Normalizar la imagen para que los valores estén en el rango [0, 255]
            imagen_normalizada_jpg = (self.imagen_normalizada * 255).astype(np.uint8)

            # Guardar la imagen normalizada en formato JPG utilizando OpenCV
            cv2.imwrite(os.path.join('resultados_algoritmos', 'imagen_segmentada.jpg'), imagen_normalizada_jpg)

    def orientar_crestas(self) -> None:
        sze = int(np.fix(6 * self.sigma_gradiente))
        if sze % 2 == 0:
            sze += 1

        gauss = cv2.getGaussianKernel(sze, self.sigma_gradiente)
        filtro_gauss = gauss @ gauss.T
        filtro_grad_y, filtro_grad_x = np.gradient(filtro_gauss)
        gradiente_x = signal.convolve2d(self.imagen_normalizada, filtro_grad_x, mode="same")
        gradiente_y = signal.convolve2d(self.imagen_normalizada, filtro_grad_y, mode="same")
        grad_x2 = gradiente_x ** 2
        grad_y2 = gradiente_y ** 2
        grad_xy = gradiente_x * gradiente_y

        sze = int(np.fix(6 * self.sigma_bloque))
        gauss = cv2.getGaussianKernel(sze, self.sigma_bloque)
        filtro_gauss = gauss @ gauss.T
        grad_x2 = ndimage.convolve(grad_x2, filtro_gauss)
        grad_y2 = ndimage.convolve(grad_y2, filtro_gauss)
        grad_xy = 2 * ndimage.convolve(grad_xy, filtro_gauss)

        denominador = np.sqrt(grad_xy ** 2 + (grad_x2 - grad_y2) ** 2) + np.finfo(float).eps
        sin_2_theta = grad_xy / denominador
        cos_2_theta = (grad_x2 - grad_y2) / denominador

        if self.sigma_suavizado_orientacion:
            sze = int(np.fix(6 * self.sigma_suavizado_orientacion))
            if sze % 2 == 0:
                sze += 1
            gauss = cv2.getGaussianKernel(sze, self.sigma_suavizado_orientacion)
            filtro_gauss = gauss @ gauss.T
            cos_2_theta = ndimage.convolve(cos_2_theta, filtro_gauss)
            sin_2_theta = ndimage.convolve(sin_2_theta, filtro_gauss)

        self.orientacion_imagen = np.pi / 2 + np.arctan2(sin_2_theta, cos_2_theta) / 2

 
        orientacion_normalizada = (self.orientacion_imagen * 255).astype(np.uint8)

     
        cv2.imwrite(os.path.join('resultados_algoritmos', 'imagen_orientacion.jpg'), orientacion_normalizada)

    def frecuencia_crestas(self):
        filas, columnas = self.imagen_normalizada.shape
        frecuencia = np.zeros((filas, columnas))

        for i in range(0, filas - self.tam_bloque_frecuencia_crestas, self.tam_bloque_frecuencia_crestas):
            for j in range(0, columnas - self.tam_bloque_frecuencia_crestas, self.tam_bloque_frecuencia_crestas):
                bloque_imagen = self.imagen_normalizada[i:i + self.tam_bloque_frecuencia_crestas, j:j + self.tam_bloque_frecuencia_crestas]
                bloque_orientacion = self.orientacion_imagen[i:i + self.tam_bloque_frecuencia_crestas, j:j + self.tam_bloque_frecuencia_crestas]
                frecuencia[i:i + self.tam_bloque_frecuencia_crestas, j:j + self.tam_bloque_frecuencia_crestas] = self.calcular_frecuencia(bloque_imagen, bloque_orientacion)

        self.frecuencia = frecuencia * self.mascara
        frecuencia_1d = np.reshape(self.frecuencia, (1, filas * columnas))
        ind = np.where(frecuencia_1d > 0)
        ind = np.array(ind)
        ind = ind[1, :]
        elementos_no_cero_en_frecuencia = frecuencia_1d[0][ind]
        self.frecuencia_media = np.mean(elementos_no_cero_en_frecuencia)
        self.frecuencia_mediana = np.median(elementos_no_cero_en_frecuencia)
        self.frecuencia = self.frecuencia_media * self.mascara

    def calcular_frecuencia(self, bloque_imagen: np.ndarray, bloque_orientacion: np.ndarray) -> np.ndarray:
        filas, _ = np.shape(bloque_imagen)
        cos_orientacion = np.mean(np.cos(2 * bloque_orientacion))
        sin_orientacion = np.mean(np.sin(2 * bloque_orientacion))
        orientacion = math.atan2(sin_orientacion, cos_orientacion) / 2
        imagen_rotada = scipy.ndimage.rotate(bloque_imagen, orientacion / np.pi * 180 + 90, axes=(1, 0), reshape=False, order=3, mode="nearest")
        tam_corte = int(np.fix(filas / np.sqrt(2)))
        desplazamiento = int(np.fix((filas - tam_corte) / 2))
        imagen_rotada = imagen_rotada[desplazamiento:desplazamiento + tam_corte, desplazamiento:desplazamiento + tam_corte]
        proyeccion = np.sum(imagen_rotada, axis=0)
        dilatacion = scipy.ndimage.grey_dilation(proyeccion, self.tam_ventana_frecuencia_crestas, structure=np.ones(self.tam_ventana_frecuencia_crestas))
        temp = np.abs(dilatacion - proyeccion)
        umbral_pico = 2
        maximos = (temp < umbral_pico) & (proyeccion > np.mean(proyeccion))
        maximos_indices = np.where(maximos)
        _, columnas_maximos_indices = np.shape(maximos_indices)

        if columnas_maximos_indices < 2:
            return np.zeros(bloque_imagen.shape)
        numero_picos = columnas_maximos_indices
        longitud_onda = (maximos_indices[0][columnas_maximos_indices - 1] - maximos_indices[0][0]) / (numero_picos - 1)
        if self.longitud_de_onda_minima <= longitud_onda <= self.longitud_de_onda_maxima:
            return 1 / np.double(longitud_onda) * np.ones(bloque_imagen.shape)
        return np.zeros(bloque_imagen.shape)

    def filtrar_crestas(self):
        imagen_normalizada = np.double(self.imagen_normalizada)
        filas, columnas = imagen_normalizada.shape
        nueva_imagen = np.zeros((filas, columnas))
        frecuencia_1d = np.reshape(self.frecuencia, (1, filas * columnas))
        ind = np.where(frecuencia_1d > 0)
        ind = np.array(ind)
        ind = ind[1, :]
        elementos_no_cero_en_frecuencia = frecuencia_1d[0][ind]
        elementos_no_cero_en_frecuencia = np.double(np.round((elementos_no_cero_en_frecuencia * 100))) / 100
        frecuencias_unicas = np.unique(elementos_no_cero_en_frecuencia)
        sigmax = 1 / frecuencias_unicas[0] * self.factor_escala_relativo_x
        sigmay = 1 / frecuencias_unicas[0] * self.factor_escala_relativo_y
        sze = int(np.round(3 * np.max([sigmax, sigmay])))
        mesh_x, mesh_y = np.meshgrid(np.linspace(-sze, sze, (2 * sze + 1)), np.linspace(-sze, sze, (2 * sze + 1)))
        filtro_referencia = np.exp(-(((np.power(mesh_x, 2)) / (sigmax * sigmax) + (np.power(mesh_y, 2)) / (sigmay * sigmay)))) * np.cos(
            2 * np.pi * frecuencias_unicas[0] * mesh_x
        )
        filas_filtro, columnas_filtro = filtro_referencia.shape
        rango_angulos = int(180 / self.incremento_angulo)
        filtro_gabor = np.array(np.zeros((rango_angulos, filas_filtro, columnas_filtro)))

        for indice_filtro in range(0, rango_angulos):
            filtro_rotado = scipy.ndimage.rotate(filtro_referencia, -(indice_filtro * self.incremento_angulo + 90), reshape=False)
            filtro_gabor[indice_filtro] = filtro_rotado

        maxsze = int(sze)
        temp = self.frecuencia > 0
        filas_validas, columnas_validas = np.where(temp)
        temp1 = filas_validas > maxsze
        temp2 = filas_validas < filas - maxsze
        temp3 = columnas_validas > maxsze
        temp4 = columnas_validas < columnas - maxsze
        final_temp = temp1 & temp2 & temp3 & temp4
        finalind = np.where(final_temp)
        maxima_cantidad_orientaciones = np.round(180 / self.incremento_angulo)
        orientacion_indice = np.round(self.orientacion_imagen / np.pi * 180 / self.incremento_angulo)

        for i in range(0, filas):
            for j in range(0, columnas):
                if orientacion_indice[i][j] < 1:
                    orientacion_indice[i][j] += maxima_cantidad_orientaciones
                if orientacion_indice[i][j] > maxima_cantidad_orientaciones:
                    orientacion_indice[i][j] -= maxima_cantidad_orientaciones

        _, columnas_finalind = np.shape(finalind)
        sze = int(sze)
        for k in range(0, columnas_finalind):
            fila_actual = filas_validas[finalind[0][k]]
            columna_actual = columnas_validas[finalind[0][k]]
            bloque_imagen = imagen_normalizada[fila_actual - sze:fila_actual + sze + 1, columna_actual - sze:columna_actual + sze + 1]
            nueva_imagen[fila_actual][columna_actual] = np.sum(bloque_imagen * filtro_gabor[int(orientacion_indice[fila_actual][columna_actual]) - 1])

        self.imagen_binaria = nueva_imagen < self.umbral_filtro_crestas

    def mejorar(self, img: np.ndarray, redimensionar: bool = True) -> np.ndarray:

        if redimensionar:
            filas, columnas = np.shape(img)
            relacion_aspecto = np.double(filas) / np.double(columnas)
            nuevas_columnas_redimensionar = self.nuevas_filas_redimensionar / relacion_aspecto
            img = cv2.resize(img, (int(nuevas_columnas_redimensionar), int(self.nuevas_filas_redimensionar)))

        imagen_normalizada = self.normalizar(img)
        self.segmentar_crestas(imagen_normalizada)
        self.orientar_crestas()  
        self.frecuencia_crestas()  
        self.filtrar_crestas() 
        return self.imagen_binaria

#######################################################################


def agregar_implicado(request):
    if request.method == 'POST':        
        nuevo_implicado = Implicado(
            caso=request.POST.get('caso_seleccionar'),
            nombres=request.POST.get('implicado_nombres'),
            apellido_paterno=request.POST.get('implicado_apellido_paterno'),
            apellido_materno=request.POST.get('implicado_apellido_materno'),
            carnet_identidad=request.POST.get('implicado_carnet'),
            expedido_en=request.POST.get('implicado_expedido'),
            fecha_nacimiento=request.POST.get('implicado_fecha_nacimiento'),
            lugar_nacimiento=request.POST.get('implicado_lugar_nacimiento'),
            direccion=request.POST.get('implicado_direccion'),
            nacionalidad=request.POST.get('implicado_nacionalidad'),
            ocupacion=request.POST.get('implicado_ocupacion'),
            color_pelo=request.POST.get('implicado_color_pelo'),
            color_piel=request.POST.get('implicado_color_piel'),
            peso_kg=request.POST.get('implicado_peso'),
            altura_m=request.POST.get('implicado_altura'),
            activo=True  
        )
        nuevo_implicado.save()

        dedos = ['pulgar', 'indice', 'medio', 'anular', 'menique']
        for dedo in dedos:
            huella = request.FILES.get(f'huella_{dedo}')
            if huella:
                try:
                    nparr = np.fromstring(huella.read(), np.uint8)
                    img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)
                    procesada_img = PreprocesamientoHuellas().mejorar(img)
                    procesada_img = (procesada_img * 255).astype(np.uint8)
                    procesada_img2 = 255 - procesada_img
                    datos, imagen_con_minucias = extraer_caracteristicas_minuciosas(procesada_img)
                    terminaciones = [(x, y) for x, y in zip(datos['terminoX'], datos['terminoY'])]
                    bifurcaciones = [(x, y) for x, y in zip(datos['bifurX'], datos['bifurY'])]
                    coordenadas_minucias = terminaciones + bifurcaciones
                    imagen_con_minucias = 255 - imagen_con_minucias

                    classification = clasificar_huella(imagen_con_minucias)

                    imag = cv2.cvtColor(imagen_con_minucias, cv2.COLOR_BGR2GRAY)
                    nucleus_coords = buscar_nucleo_esqueleto(imag)

                    huella_instancia = Huellas()
                    huella_instancia.persona = nuevo_implicado.carnet_identidad  # Asignando el carnet_identidad del implicado
                    huella_instancia.caso = nuevo_implicado.caso
                    huella_instancia.original_image.save(huella.name, huella, save=False)
                    huella_instancia.procesada_image.save(huella.name, ContentFile(cv2.imencode('.png', procesada_img2)[1]), save=False)

                    huella_instancia.tipo = classification
                    huella_instancia.minutiae_image.save(huella.name, ContentFile(cv2.imencode('.png', imagen_con_minucias)[1]), save=False)
                    huella_instancia.save()

                    if nucleus_coords:
                        csv_bytes_io = BytesIO()
                        csv_filename = f"{huella.name.split('.')[0]}_minucias.csv"
                        csv_string_io = StringIO()
                        writer = csv.writer(csv_string_io)
                        writer.writerow(["X", "Y", "Distancia", "Angulo"])
                        for point in coordenadas_minucias:
                            distance, angle = calculate_distance_angle(nucleus_coords, point)
                            writer.writerow([point[0], point[1], distance, angle])
                        csv_string_io.seek(0)
                        huella_instancia.minutiae_csv.save(csv_filename, ContentFile(csv_string_io.getvalue().encode('utf-8')), save=True)
                        huella_instancia.save()

                except Exception as e:
                    print(f"Error al procesar la imagen: {str(e)}")
        
        messages.success(request, 'Implicado agregado exitosamente con huellas procesadas.')
        return redirect('listar_huella')
    else:
        casos_activos = Casos.objects.filter(activate=True).order_by('-f_registro')
        return render(request, "casos/agregar_implicado.html", {'casos_activos': casos_activos})

def find_nucleus(skeleton_image):
    skeleton_image = cv2.bitwise_not(skeleton_image)
    
    contours, _ = cv2.findContours(skeleton_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not contours:
        return None
    
    nucleus_contour = max(contours, key=cv2.contourArea)

    M = cv2.moments(nucleus_contour)
    nucleus_center = (int(M["m10"] / M["m00"]), int(M["m01"] / M["m00"]))
    
    return nucleus_center

def calculate_distance_angle(nucleo, punto):
    x1, y1 = nucleo
    x2, y2 = punto
    distance = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
    angle = math.degrees(math.atan2(y2 - y1, x2 - x1))
    return distance, angle



def detect_minutiae(skeleton):
    corners = cv2.cornerHarris(skeleton, 9, 9, 0.1)
    corners = cv2.dilate(corners, None)
    ret, corners = cv2.threshold(corners, 0.01 * corners.max(), 255, 0)
    corners = np.uint8(corners)
    ret, labels, stats, centroids = cv2.connectedComponentsWithStats(corners)
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.01)
    corners = cv2.cornerSubPix(skeleton, np.float32(centroids), (5, 5), (-1, -1), criteria)
    minutiae_points = [(int(c[0]), int(c[1])) for c in corners[1:]]  
    return minutiae_points

def marcar_minutiae(imagen, minutiae_points):
    imagen_np = np.array(imagen)
    if len(imagen_np.shape) == 2:
        imagen_np = cv2.cvtColor(imagen_np, cv2.COLOR_GRAY2BGR)

    for point in minutiae_points:
        cv2.circle(imagen_np, point, 5, (255, 0, 0), 1)

    imagen_procesada = Image.fromarray(imagen_np)
    return imagen_procesada

def leer_datos_minucia(csv_path):
    
    datos_minucia = []
    with open(csv_path, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            datos_minucia.append({'distancia': float(row['Distancia']), 'angulo': float(row['Angulo'])})
    return datos_minucia

def calcular_similitud(datos_huella_subida, datos_huella_guardada):
    coincidencias = 0
    umbral_distancia = 5  
    umbral_angulo = 10 
    for dato_subida in datos_huella_subida:
        for dato_guardado in datos_huella_guardada:
            if abs(dato_subida['distancia'] - dato_guardado['distancia']) < umbral_distancia and abs(dato_subida['angulo'] - dato_guardado['angulo']) < umbral_angulo:
                coincidencias += 1
                break  
    return coincidencias

def preprocesar_huella(img):
    aplicar_algoritmo_imagen = PreprocesamientoHuellas()  
    imagen_procesada = aplicar_algoritmo_imagen.mejorar(img)  
    imagen_procesada = (imagen_procesada * 255).astype(np.uint8)
    datos,imagen_con_minucias =  extraer_caracteristicas_minuciosas(imagen_procesada)
    terminaciones = [(x, y) for x, y in zip(datos['terminoX'], datos['terminoY'])]
    bifurcaciones = [(x, y) for x, y in zip(datos['bifurX'], datos['bifurY'])]
    coordenadas_minucias = terminaciones + bifurcaciones
    imagen_con_minucias=255-imagen_con_minucias
    return imagen_con_minucias,coordenadas_minucias




def procesar_imagen(request):
    if request.method == 'POST' and request.FILES.get('imagen'):
        imagen = request.FILES['imagen']
        nparr = np.fromstring(imagen.read(), np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)
        imagen_procesada,minucias= preprocesar_huella(img)
        # print(minucias)
        imag = cv2.cvtColor(imagen_procesada, cv2.COLOR_BGR2GRAY)
        nucleus_coords = buscar_nucleo_esqueleto(imag)
        minutiae_data = []
        if nucleus_coords:
            for point in minucias:
                distance, angle = calculate_distance_angle(nucleus_coords, point)
                minutiae_data.append({'point': point, 'distance': distance, 'angle': angle})
                print(minutiae_data)
                
        resultados_comparacion = comparar_con_huellas_guardadas(minutiae_data)

        if imagen_procesada.dtype != np.uint8:
            imagen_procesada = imagen_procesada.astype(np.uint8)
        _, buffer = cv2.imencode('.png', imagen_procesada)
        imagen_base64 = base64.b64encode(buffer)
        classification = clasificar_huella(img)
        response_data = {
                'imagen': imagen_base64.decode(),
                'classificacion': classification,
                'nucleo': nucleus_coords,
                'puntos': len(minucias),
                'resultados_comparacion': resultados_comparacion
        }
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)



def cotejar_huella(request):
     return render(request, "huellas/comparar_huella.html")

def comparar_huella(request):
    return render(request, "huellas/comparar_huella.html")


def obtener_imagenes_huella(request, huella_id):
    try:
        huella = Huellas.objects.get(pk=huella_id)
        
        urlOriginal = huella.original_image.url if huella.original_image else ''
        urlProcesada = huella.procesada_image.url if huella.procesada_image else ''
        urlMinucias = huella.minutiae_image.url if huella.minutiae_image else ''
        
        if huella.persona:
            try:
                implicado = Implicado.objects.get(carnet_identidad=huella.persona)
                nombre_completo = f"{implicado.nombres} {implicado.apellido_paterno} {implicado.apellido_materno}"
            except Implicado.DoesNotExist:
                nombre_completo = "No encontrado"
        else:
            nombre_completo = "No especificado"

        return JsonResponse({
            'urlOriginal': urlOriginal,
            'urlProcesada': urlProcesada,
            'urlMinucias': urlMinucias,
            'nombrePersona': nombre_completo
        })
    except Huellas.DoesNotExist:
        return JsonResponse({'error': 'Huella no encontrada'}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Una o más imágenes no se pudieron encontrar'}, status=404)

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Table, TableStyle, Spacer
from reportlab.lib.units import inch
from reportlab.lib.colors import Color
from django.http import HttpResponse
from urllib.parse import unquote
import datetime

from reportlab.lib.enums import TA_JUSTIFY
def generar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_de_coincidencia.pdf"'
    imagen_original = unquote(request.GET.get('imagen_original', ''))
    imagen_procesada = unquote(request.GET.get('imagen_procesada', ''))
    nivel_coincidencia = request.GET.get('nivel_coincidencia', '')
    persona_identificada = request.GET.get('persona_identificada', '')
    
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=10)
    styles = getSampleStyleSheet()
    encabezado_style = ParagraphStyle(name='encabezado', fontSize=12)
    contenido = []
    centered_style = ParagraphStyle(name='centered', alignment=1, parent=styles['Normal'])
    logo = "media/logo.jpg"  # Asegúrate de poner la ruta correcta del logo
    verde_oscuro = Color(0, 0.5, 0)  # Verde oscuro
    encabezado_izquierda = [
        [
            Image(logo, width=90, height=90),
            Paragraph("<b>INFORME DE RESULTADOS</b><br/><br/><b>División:</b> Primera<br/><b>Sección:</b> 2da Sección<br/><b>Perito Responsable:</b>", encabezado_style)
        ]
    ]
    encabezado_derecha = [
        ["R.U.P.", "", "TDJ", ""],
        ["Div", "", "SP", ""],
        ["MP", "", "Página", "1"],
        [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    tabla_encabezado_izquierda = Table(encabezado_izquierda, colWidths=[100, 150])
    tabla_encabezado_izquierda.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    tabla_encabezado_derecha = Table(encabezado_derecha, colWidths=[40, 60, 40, 60])
    tabla_encabezado_derecha.setStyle(TableStyle([
        ('SPAN', (0, 3), (-1, 3)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
    ]))
    encabezado_completo = Table(
        [
            [tabla_encabezado_izquierda, tabla_encabezado_derecha]
        ],
        colWidths=[330, 310]
    )
    encabezado_completo.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (1, 1), 'TOP'),
    ]))
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 10))
    contenido.append(encabezado_completo)
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 20))  
    contenido.append(Paragraph("<b>INFORME PERICIAL</b>", styles['Title']))
    contenido.append(Paragraph("<b>D A C T I L O S C O P I C O</b>", styles['Title']))
    contenido.append(Spacer(1, 20))
    contenido.append(Paragraph("<b>I.- ANTECEDENTES.</b>", styles['Heading2']))
    contenido_style = ParagraphStyle(
    name='contenido',
    alignment=TA_JUSTIFY,
    fontSize=12,
    )
    contenido.append(Paragraph("Conforme al oficio CITE: No. 0WWW/VV/DCM, de fecha AA de enero de 2024, emitido por el DIRECTOR NACIONAL O.C.N. INTERPOL, se procedió a realizar la búsqueda y cotejo de la imagen ingresada con la base de datos en busca de coincidencias. Como resultado, se obtuvo una coincidencia con la persona registrada como: Sr. "+persona_identificada, contenido_style))
    contenido.append(Paragraph("<br/><b>II.- OBJETO DE LA PERICIA.</b>", styles['Heading2']))
    contenido.append(Paragraph("Realizar la COMPARACION DACTILOSCOPICA, de las impresiones digitales cuestionadas estampadas en la ficha decadactilar, con relación a las impresiones digitales de comparación obrantes en la Tarjeta Prontuario No. 8332129 a nombre de Casimiro La Luz", contenido_style))
    contenido.append(Paragraph("<br/><b>III.- BASE DE LA PERICIA.</b>", styles['Heading2']))
    contenido.append(Paragraph("Antes de empezar la tarea pericial misma, es necesario señalar los fundamentos técnicos científicos en los que se basan las pericias en dactiloscopia, tendientes a determinar la identidad fisica humana.La dactiloscopia se apoya en los postulados de la perennidad, inmutabilidad y la variedad infinita.<br/><br/>PERENNIDAD: Los dibujos digitales se forman a partir del séptimo y octava semana de la vida intra-uterina y desde ese momento el dibujo dactilar es perenne a través de toda la existencia del ser humano. Su destrucción solo se produce con la desintegración de la piel, luego de transcurrir un largo periodo después de la muerte y cuando sobrevienen los fenómenos de putrefacción cadavérica.<br/><br/>INMUTABLILIDAD: Es decir, los dibujos dactilares no cambian jamás, al extremo de que ni por propia voluntad, ni por circunstancias patológicas o enfermedades y en aquellos casos en que aparece deterioro de la epidermis, éste es transitorio, ya que una vez cesada la causa, el dibujo reaparece exactamente igual a su forma original. Además el dibujo dactilar no desaparece mientras no haya sufrido una lesión o quemadura que afecta profundamente la dermis.<br/><br/>", contenido_style))
    contenido.append(Spacer(1, 10))
    verde_oscuro = Color(0, 0.5, 0)  # Verde oscuro
    encabezado_izquierda = [
        [
            Image(logo, width=90, height=90),
            Paragraph("<b>INFORME DE RESULTADOS</b><br/><br/><b>División:</b> Primera<br/><b>Sección:</b> 2da Sección<br/><b>Perito Responsable:</b>", encabezado_style)
        ]
    ]
    encabezado_derecha = [
        ["R.U.P.", "", "TDJ", ""],
        ["Div", "", "SP", ""],
        ["MP", "", "Página", "2"],
        [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    tabla_encabezado_izquierda = Table(encabezado_izquierda, colWidths=[100, 150])
    tabla_encabezado_izquierda.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    tabla_encabezado_derecha = Table(encabezado_derecha, colWidths=[40, 60, 40, 60])
    tabla_encabezado_derecha.setStyle(TableStyle([
        ('SPAN', (0, 3), (-1, 3)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
    ]))
    encabezado_completo = Table(
        [
            [tabla_encabezado_izquierda, tabla_encabezado_derecha]
        ],
        colWidths=[330, 310]
    )
    encabezado_completo.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (1, 1), 'TOP'),
    ]))
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 10))
    contenido.append(encabezado_completo)
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 30))
    contenido.append(Paragraph("<br/>VARIEDAD INFINITA: En lineas generales, este principio asegura que dada la gran cantidad de puntos característicos que existen en una impresión digital, las combinaciones que se pueden realizar con ellos es de tal magnitud,que matemáticamente no existe la posibilidad de que se repita una impresión digital identica en dos seres humanos, es más, en diferentes dedos de un mismo ser, en miles de años.<br/><br/>", contenido_style))
    contenido.append(Paragraph("<b>IV.- OPERACIONES REALIZADAS</b>", styles['Heading2']))
    contenido.append(Paragraph("A los efectos de una mejor interpretación de todo cuanto se vaya a exponer en el transcurso del presente informe, el suscrito cree conveniente exponer el siguiente esquema de trabajo <br/><br/>Con la ayuda de instrumental óptico y fuentes lumínicas necesarios para este tipo de pericias; además de equipo computacional con software para el efecto,<br/><br/>se evidencia lo siguiente:<br></br><br/>a) Las impresiones dactilares cuestionadas, obrantes en la ficha decadactilar, reúnen condiciones suficientes de nitidez e integridad para cotejo de carácter dactiloscópico.<br/><br/>b) Las impresión dactilares de comparación, obrantes, en la tarjera prontuario, obtenida como evidencia reúne condiciones suficientes de nitidez e integridad para cotejo dactiloscópico.", contenido_style))
    contenido.append(Paragraph("<br/><b>V.- COTEJO DE HUELLAS E IMPRESIONES DACTILARES</b>", styles['Heading2']))
    contenido.append(Paragraph("Realizada la comparacion formal de las impresiones ingresando dichas al software LUKANA para obtener similitudes en cuanto a las huellas cuestionadas y encontradas con el apoyo de herramientas externas se presentaran los resultados obtenidos.", contenido_style))
    # contenido.append(Paragraph("<br/><b>5.1.- Impresiones dactilares a escalas grises</b>", contenido_style))
    # contenido.append(Paragraph("<br/>A continuacion se presentan las huellas dactilares cuestionada con la huella encontrada a escalas grises para su uso posterior para verificacion de puntos y caracteristicas encontrados:", contenido_style))
    # data = [
    #     [
    #         Paragraph("<b>IMPRESIÓN DIGITAL CUESTIONADA</b>", centered_style),
    #         Paragraph("<b>IMPRESIÓN DIGITAL ENCONTRADA</b>", centered_style)
    #     ],
    #     [
    #         Image(imagen_original, width=200, height=250),
    #         Image(imagen_procesada, width=200, height=250)
    #     ]
    # ]
    # for row in data:
    #     for cell in row:
    #         if isinstance(cell, Image):
    #             cell.drawWidth = cell._width
    #             cell.drawHeight = cell._height

    # tabla_imagenes = Table(data, colWidths=[280, 280])
    # tabla_imagenes.setStyle(TableStyle([
    #     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    # ]))
    # contenido.append(Spacer(1, 20))
    # contenido.append(tabla_imagenes)
    contenido.append(Paragraph("<br/><br/><b>5.1.- Esqueleto de impresiones dactilares con puntos caracteristicos hallados</b>", contenido_style))
    contenido.append(Paragraph("<br/>A continuación se detalla el análisis dactiloscópico completo, presentando el esqueleto detallado de las huellas dactilares cuestionadas, acompañado de las huellas encontradas por el sistema, resaltando los puntos y características distintivas detectadas durante el proceso de comparación:", contenido_style))
    contenido.append(Spacer(1, 45))
    verde_oscuro = Color(0, 0.5, 0)  # Verde oscuro
    encabezado_izquierda = [
        [
            Image(logo, width=90, height=90),
            Paragraph("<b>INFORME DE RESULTADOS</b><br/><br/><b>División:</b> Primera<br/><b>Sección:</b> 2da Sección<br/><b>Perito Responsable:</b>", encabezado_style)
        ]
    ]
    encabezado_derecha = [
        ["R.U.P.", "", "TDJ", ""],
        ["Div", "", "SP", ""],
        ["MP", "", "Página", "3"],
        [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    tabla_encabezado_izquierda = Table(encabezado_izquierda, colWidths=[100, 150])
    tabla_encabezado_izquierda.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    tabla_encabezado_derecha = Table(encabezado_derecha, colWidths=[40, 60, 40, 60])
    tabla_encabezado_derecha.setStyle(TableStyle([
        ('SPAN', (0, 3), (-1, 3)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
    ]))
    encabezado_completo = Table(
        [
            [tabla_encabezado_izquierda, tabla_encabezado_derecha]
        ],
        colWidths=[330, 310]
    )
    encabezado_completo.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (1, 1), 'TOP'),
    ]))
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 10))
    contenido.append(encabezado_completo)
    contenido.append(Table(
        [[Paragraph("")]],
        colWidths=[540],
        style=[('LINEBELOW', (0, 0), (-1, 0), 8, verde_oscuro)]
    ))
    contenido.append(Spacer(1, 30))
    data = [
        [
            Paragraph("<b>IMPRESIÓN DIGITAL CUESTIONADA</b>", centered_style),
            Paragraph("<b>IMPRESIÓN DIGITAL ENCONTRADA</b>", centered_style)
        ],
        [
            Image(imagen_original, width=200, height=250),
            Image(imagen_procesada, width=200, height=250)
        ]
    ]
    for row in data:
        for cell in row:
            if isinstance(cell, Image):
                cell.drawWidth = cell._width
                cell.drawHeight = cell._height
    tabla_imagenes = Table(data, colWidths=[280, 280])
    tabla_imagenes.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    contenido.append(Spacer(1, 20))
    contenido.append(tabla_imagenes)
    contenido.append(Paragraph("<b>VI.- CONCLUSIONES</b>", styles['Heading2']))
    contenido.append(Paragraph("Las impresiones dactilares cuestionadas obrantes en la ficha PRONTUARIA sin identificación, CORRESPONDEN A LOS DACTILOGRAMAS DE COMPARACION DE: "+persona_identificada, contenido_style))
    contenido.append(Spacer(1, 20))
    contenido.append(Paragraph(f"<b>Nivel de Coincidencia:</b> {nivel_coincidencia} % de coincidencia", styles['Heading2']))
    contenido.append(Paragraph("<b>Persona Identificada:</b> " + persona_identificada, styles['Normal']))
    # contenido.append(Paragraph("<b>Nivel de Coincidencia:</b> " + nivel_coincidencia, styles['Normal']))
    
    doc.build(contenido)
    return response